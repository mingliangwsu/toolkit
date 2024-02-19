#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Aug 18 14:25:37 2023

@author: liuming
"""

import openpyxl
import sys

if len(sys.argv) <= 1:
    print("Usage:" + sys.argv[0] + "<1input> <2output_dir>\n")
    sys.exit(0)

# Load the Excel workbook
workbook = openpyxl.load_workbook(sys.argv[1], data_only=True)  # data_only=True to read values, not formulas

# Define the output directory path
output_directory = sys.argv[2]

# Iterate through all sheets in the workbook
for sheet_name in workbook.sheetnames:
    print(sheet_name)
    worksheet = workbook[sheet_name]  # Get the sheet by its name
    
    # Define the output text file path for this sheet
    output_path = output_directory + sheet_name + '.txt'
    
    # Open the output text file for writing
    with open(output_path, 'w') as output_file:
        for row in worksheet.iter_rows(values_only=True):  # Iterate through rows, fetching values only
            #tab_delimited_row = '\t'.join(str(cell) for cell in row)  # Combine row values with tabs
            tab_delimited_row = '\t'.join(str(cell) if cell is not None else '' for cell in row)  # Replace None with ''
            output_file.write(tab_delimited_row + '\n')  # Write the row to the text file

print("Export completed.")